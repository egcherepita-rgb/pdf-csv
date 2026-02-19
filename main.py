import io
import os
import re
import csv
import time
import json
import threading
from uuid import uuid4
from collections import OrderedDict
from typing import List, Tuple, Dict, Optional

import fitz  # PyMuPDF
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import Response, HTMLResponse, FileResponse

try:
    import openpyxl  # requires openpyxl in requirements.txt
except Exception:
    openpyxl = None


app = FastAPI(
    title="PDF → CSV (артикул / наименование / всего / категория)",
    version="4.2.0",
)

# -------------------------
# Regex
# -------------------------
RX_SIZE = re.compile(r"\b\d{2,}[xх×]\d{2,}(?:[xх×]\d{1,})?\b", re.IGNORECASE)
RX_MM = re.compile(r"мм", re.IGNORECASE)
RX_WEIGHT = re.compile(r"\b\d+(?:[.,]\d+)?\s*кг\.?\b", re.IGNORECASE)

RX_MONEY_LINE = re.compile(r"^\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?\s*₽$")
RX_INT = re.compile(r"^\d+$")
RX_ANY_RUB = re.compile(r"₽")

# В одной строке: "... 450 ₽ 2 900 ₽"
RX_PRICE_QTY_SUM = re.compile(
    r"(?P<price>\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*₽\s+"
    r"(?P<qty>\d{1,4})\s+"
    r"(?P<sum>\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*₽",
    re.IGNORECASE,
)

# Размеры где угодно в строке: "30x10 мм", "600×300 мм" и т.п.
RX_DIMS_ANYWHERE = re.compile(
    r"\s*\d{1,4}[xх×]\d{1,4}(?:[xх×]\d{1,5})?\s*мм\.?\s*",
    re.IGNORECASE,
)

# -------------------------
# Helpers
# -------------------------
def normalize_space(s: str) -> str:
    s = (s or "").replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_key(name: str) -> str:
    s = normalize_space(name).lower()
    s = s.replace("×", "x").replace("х", "x")
    s = RX_DIMS_ANYWHERE.sub(" ", s)
    s = normalize_space(s)
    return s


def strip_dims_anywhere(name: str) -> str:
    name = normalize_space(name)
    name2 = RX_DIMS_ANYWHERE.sub(" ", name)
    return normalize_space(name2)


# -------------------------
# Артикулы (Art.xlsx)
# -------------------------
def load_article_map() -> Tuple[Dict[str, str], str]:
    if openpyxl is None:
        return {}, "openpyxl_not_installed"

    path = os.getenv("ART_XLSX_PATH", "Art.xlsx")
    if not os.path.exists(path):
        return {}, f"file_not_found:{path}"

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return {}, f"cannot_open:{e}"

    header = [normalize_space(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    товар_col = 1
    арт_col = 2
    for idx, h in enumerate(header, start=1):
        if h.lower() == "товар":
            товар_col = idx
        if h.lower() == "артикул":
            арт_col = idx

    m: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        товар = ws.cell(r, товар_col).value
        арт = ws.cell(r, арт_col).value
        if not товар or not арт:
            continue
        товар_s = normalize_space(str(товар))
        арт_s = normalize_space(str(арт))
        if not товар_s or not арт_s:
            continue
        m[normalize_key(товар_s)] = арт_s

    return m, "ok"


ARTICLE_MAP, ARTICLE_MAP_STATUS = load_article_map()

CATEGORY_VALUE = 2

# -------------------------
# Счетчик конвертаций (простое файловое хранилище)
# -------------------------
from threading import Lock

COUNTER_FILE = os.getenv("COUNTER_FILE", "conversions.count")
_counter_lock = Lock()


def _read_counter() -> int:
    try:
        with open(COUNTER_FILE, "r", encoding="utf-8") as f:
            return int((f.read() or "0").strip())
    except Exception:
        return 0


def _write_counter(v: int) -> None:
    tmp = COUNTER_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(str(v))
    os.replace(tmp, COUNTER_FILE)


def increment_counter() -> int:
    with _counter_lock:
        v = _read_counter() + 1
        _write_counter(v)
        return v


def get_counter() -> int:
    return _read_counter()


# -------------------------
# Async jobs (progress) — FILE storage (OK with gunicorn -w 2)
# -------------------------
JOB_DIR = os.getenv("JOB_DIR", "/tmp/pdf2csv_jobs")
JOB_TTL_SEC = int(os.getenv("JOB_TTL_SEC", str(30 * 60)))  # seconds


def _ensure_job_dir() -> None:
    os.makedirs(JOB_DIR, exist_ok=True)


def _job_json_path(job_id: str) -> str:
    return os.path.join(JOB_DIR, f"{job_id}.json")


def _job_result_path(job_id: str) -> str:
    return os.path.join(JOB_DIR, f"{job_id}.csv")


def _write_json_atomic(path: str, data: dict) -> None:
    _ensure_job_dir()
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, path)


def _read_json(path: str) -> Optional[dict]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _cleanup_jobs() -> None:
    _ensure_job_dir()
    now = time.time()
    try:
        for name in os.listdir(JOB_DIR):
            if not (name.endswith(".json") or name.endswith(".csv")):
                continue
            p = os.path.join(JOB_DIR, name)
            try:
                if now - os.path.getmtime(p) > JOB_TTL_SEC:
                    show = True
                    os.remove(p)
            except Exception:
                pass
    except Exception:
        pass


def _set_job(job_id: str, **kwargs) -> None:
    _cleanup_jobs()
    path = _job_json_path(job_id)
    data = _read_json(path) or {}
    data.update(kwargs)
    _write_json_atomic(path, data)


def _get_job(job_id: str) -> Optional[dict]:
    _cleanup_jobs()
    return _read_json(_job_json_path(job_id))


def _set_job_result(job_id: str, csv_bytes: bytes) -> None:
    _cleanup_jobs()
    _ensure_job_dir()
    p = _job_result_path(job_id)
    tmp = p + ".tmp"
    with open(tmp, "wb") as f:
        f.write(csv_bytes)
    os.replace(tmp, p)


def _get_job_result(job_id: str) -> Optional[bytes]:
    p = _job_result_path(job_id)
    try:
        with open(p, "rb") as f:
            return f.read()
    except Exception:
        return None


# -------------------------
# Instruction media (optional)
# -------------------------
INSTRUCTION_VIDEO_PATH = os.getenv("INSTRUCTION_VIDEO_PATH", "instruction.mp4")

# -------------------------
# PDF parsing helpers
# -------------------------
def is_noise(line: str) -> bool:
    low = (line or "").strip().lower()
    if not low:
        return True
    if low.startswith("страница:"):
        return True
    if low.startswith("ваш проект"):
        return True
    if "проект создан" in low:
        return True
    if "развертка стены" in low:
        return True
    if "стоимость проекта" in low:
        return True
    return False


def is_totals_block(line: str) -> bool:
    low = (line or "").strip().lower()
    return (
        low.startswith("общий вес")
        or low.startswith("максимальный габарит заказа")
        or low.startswith("адрес:")
        or low.startswith("телефон:")
        or low.startswith("email")
    )


def money_to_number(line: str) -> int:
    s = normalize_space(line).replace("₽", "").strip()
    s = s.replace("\u00a0", " ")
    s = s.replace(" ", "")
    s = s.replace(",", ".")
    try:
        return int(float(s))
    except Exception:
        return -1


def is_project_total_only(line: str, prev_line: str = "") -> bool:
    """
    Не путать цену позиции с итогом проекта.
    1) Если предыдущая строка содержит "стоимость проекта" — считаем следующую сумму итогом.
    2) Если строка выглядит как "NNN ₽" и сумма >= 10 000 — считаем итогом.
    """
    if not RX_MONEY_LINE.fullmatch(normalize_space(line)):
        return False
    prev = normalize_space(prev_line).lower()
    if "стоимость проекта" in prev:
        return True
    v = money_to_number(line)
    return v >= 10000


def is_header_token(line: str) -> bool:
    low = normalize_space(line).lower().replace("–", "-").replace("—", "-")
    return low in {"фото", "товар", "габариты", "вес", "цена за шт", "кол-во", "сумма"}


def looks_like_dim_or_weight(line: str) -> bool:
    if RX_WEIGHT.search(line):
        return True
    if RX_SIZE.search(line) and RX_MM.search(line):
        return True
    return False


def looks_like_money_or_qty(line: str) -> bool:
    if RX_MONEY_LINE.fullmatch(line):
        return True
    if RX_INT.fullmatch(line):
        return True
    return False


def clean_name_from_buffer(buf: List[str]) -> str:
    filtered: List[str] = []
    for ln in buf:
        if is_noise(ln) or is_header_token(ln) or is_totals_block(ln):
            continue
        filtered.append(ln)

    while filtered and (looks_like_dim_or_weight(filtered[-1]) or looks_like_money_or_qty(filtered[-1])):
        filtered.pop()

    name = normalize_space(" ".join(filtered))
    name = re.sub(r"^Фото\s*", "", name, flags=re.IGNORECASE).strip()
    name = re.sub(r"^Товар\s*", "", name, flags=re.IGNORECASE).strip()
    name = strip_dims_anywhere(name)
    return name


# -------------------------
# Main parser
# -------------------------
def parse_items(pdf_bytes: bytes) -> Tuple[List[Tuple[str, int]], Dict]:
    """
    Поддерживает якоря:
    A) В одной строке: price ₽ qty sum ₽
    B) В разных строках: price ₽ -> qty -> sum ₽
    C) В строке с весом/размерами цена "прилипла": ... кг. 75 ₽ -> qty -> sum ₽
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    total_pages = doc.page_count

    ordered: "OrderedDict[str, int]" = OrderedDict()
    buf: List[str] = []

    stats = {
        "pages": 0,
        "total_pages": total_pages,
        "processed_pages": 0,
        "items_found": 0,
        "anchors_inline": 0,
        "anchors_multiline": 0,
        "article_map_size": len(ARTICLE_MAP),
        "article_map_status": ARTICLE_MAP_STATUS,
    }

    for page in doc:
        stats["pages"] += 1
        stats["processed_pages"] += 1

        # ---- SPEED: extract text once, skip pages without ₽
        txt = page.get_text("text") or ""
        if "₽" not in txt:
            continue

        lines = [normalize_space(x) for x in txt.splitlines()]
        lines = [x for x in lines if x]
        if not lines:
            continue
        # ----

        in_totals = False
        buf.clear()

        i = 0
        while i < len(lines):
            line = lines[i]
            prev = lines[i - 1] if i > 0 else ""

            if is_noise(line) or is_header_token(line):
                i += 1
                continue

            if is_project_total_only(line, prev_line=prev) or is_totals_block(line):
                in_totals = True
                buf.clear()
                i += 1
                continue

            if in_totals:
                i += 1
                continue

            # A) INLINE anchor
            m = RX_PRICE_QTY_SUM.search(line)
            if m:
                name = clean_name_from_buffer(buf)
                buf.clear()

                if name:
                    try:
                        qty = int(m.group("qty"))
                    except Exception:
                        qty = 0

                    if 1 <= qty <= 500:
                        ordered[name] = ordered.get(name, 0) + qty
                        stats["items_found"] += 1
                        stats["anchors_inline"] += 1

                i += 1
                continue

            # C) EMBEDDED price anchor: line contains ₽ (price embedded in weight/dims line)
            #    then next line is qty, next-next is sum money line
            if RX_ANY_RUB.search(line):
                if i + 2 < len(lines) and RX_INT.fullmatch(lines[i + 1]) and RX_MONEY_LINE.fullmatch(lines[i + 2]):
                    try:
                        qty = int(lines[i + 1])
                    except Exception:
                        qty = 0

                    if 1 <= qty <= 500:
                        # include current line so cleaner can drop weight/dims/price tail
                        name = clean_name_from_buffer(buf + [line])
                        buf.clear()

                        if name:
                            ordered[name] = ordered.get(name, 0) + qty
                            stats["items_found"] += 1
                            stats["anchors_multiline"] += 1

                    i += 3
                    continue

            # B) MULTILINE anchor: money -> qty -> money
            if RX_MONEY_LINE.fullmatch(line):
                end = min(len(lines), i + 8)

                qty_idx = None
                for j in range(i + 1, end):
                    if RX_INT.fullmatch(lines[j]):
                        q = int(lines[j])
                        if 1 <= q <= 500:
                            qty_idx = j
                            break

                if qty_idx is None:
                    buf.append(line)
                    i += 1
                    continue

                sum_idx = None
                for j in range(qty_idx + 1, end):
                    if RX_MONEY_LINE.fullmatch(lines[j]):
                        sum_idx = j
                        break

                if sum_idx is None:
                    buf.append(line)
                    i += 1
                    continue

                name = clean_name_from_buffer(buf)
                buf.clear()

                if name:
                    qty = int(lines[qty_idx])
                    ordered[name] = ordered.get(name, 0) + qty
                    stats["items_found"] += 1
                    stats["anchors_multiline"] += 1

                i = sum_idx + 1
                continue

            buf.append(line)
            i += 1

    return list(ordered.items()), stats


# -------------------------
# CSV output (Excel-friendly)
# -------------------------
def make_csv_excel_friendly(rows: List[Tuple[str, int]]) -> bytes:
    out = io.StringIO()
    writer = csv.writer(
        out,
        delimiter=";",
        quotechar='"',
        quoting=csv.QUOTE_MINIMAL,
        lineterminator="\r\n",
    )

    writer.writerow(["Артикул", "Наименование", "Всего", "Категория"])

    for name, qty in rows:
        art = ARTICLE_MAP.get(normalize_key(name), "")
        writer.writerow([art, name, qty, CATEGORY_VALUE])

    return out.getvalue().encode("utf-8-sig")  # UTF-8 BOM


# -------------------------
# UI (progress + timer)
# -------------------------
HOME_HTML = """<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>PDF → CSV</title>
  <style>
    :root { --bg:#0b0f17; --card:#121a2a; --text:#e9eefc; --muted:#a8b3d6; --border:rgba(255,255,255,.08); --btn:#4f7cff; }
    body { margin:0; font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
           background: radial-gradient(1200px 600px at 20% 10%, #18234a 0%, var(--bg) 55%); color: var(--text); }
    .wrap { min-height: 100vh; display:flex; align-items:center; justify-content:center; padding: 28px; }
    .card { width:min(900px, 100%); background: rgba(18,26,42,.92); border: 1px solid var(--border);
            border-radius: 18px; padding: 22px; box-shadow: 0 18px 60px rgba(0,0,0,.45); }
    .top { display:flex; gap:14px; align-items:center; justify-content:space-between; flex-wrap:wrap; }
    h1 { margin:0; font-size: 24px; letter-spacing: .2px; line-height: 1.15; }
    .sub { margin-top: 6px; font-size: 16px; font-weight: 700; color: var(--muted); }
    .hint { margin: 8px 0 0; color: var(--muted); font-size: 14px; }
    .badge { font-size: 12px; color: var(--muted); border: 1px solid var(--border); padding: 6px 10px; border-radius: 999px; }
    .row { margin-top: 18px; display:flex; gap: 12px; align-items:center; flex-wrap:wrap; }
    .file { display:flex; align-items:center; gap:10px; padding: 10px 12px; border: 1px dashed var(--border);
            border-radius: 14px; background: rgba(255,255,255,.02); }
    button { padding: 10px 14px; border: 0; border-radius: 14px; cursor: pointer; font-weight: 800;
             background: var(--btn); color: #0b1020; }
    button:disabled { opacity: .55; cursor:not-allowed; }
    .status { margin-top: 14px; font-size: 14px; color: var(--muted); white-space: pre-wrap; }
    .status.ok { color: #79ffa8; }
    .status.err { color: #ff7b8a; }
    .help { margin-top: 16px; }
    .helphead { display:flex; align-items:center; justify-content:space-between; gap:10px; flex-wrap:wrap; }
    .helptitle { font-weight: 700; color: var(--text); }
    .openfull { font-size: 13px; color: var(--muted); text-decoration: underline; cursor: pointer; }
    .thumb { margin-top: 10px; border: 1px solid var(--border); border-radius: 14px; overflow:hidden;
             background: rgba(255,255,255,.02); cursor: zoom-in; }
    .thumb video { display:block; width:100%; height:auto; max-height: 260px; object-fit: cover; object-position: top; }
    /* modal */
    .modal { position: fixed; inset: 0; background: rgba(0,0,0,.75); display:none; align-items:center; justify-content:center; padding: 18px; }
    .modal.open { display:flex; }
    .modalcard { width:min(1200px, 100%); background: rgba(18,26,42,.96); border: 1px solid var(--border);
                 border-radius: 18px; overflow:hidden; box-shadow: 0 30px 100px rgba(0,0,0,.6); }
    .modalbar { display:flex; align-items:center; justify-content:space-between; padding: 10px 12px; border-bottom: 1px solid var(--border); }
    .modalbar .t { color: var(--text); font-weight: 700; font-size: 14px; }
    .close { background: transparent; color: var(--muted); border: 1px solid var(--border);
             border-radius: 12px; padding: 8px 10px; cursor:pointer; font-weight: 700; }
    .modalbody { background: #0b0f17; }
    .modalbody video { display:block; width:100%; height:auto; }

    /* bottom widgets */
    .corner { position: fixed; right: 12px; bottom: 10px; font-size: 12px; color: var(--muted); opacity: .9; }
    .footer-note {
      position: fixed;
      left: 50%;
      bottom: 10px;
      transform: translateX(-50%);
      font-size: 13px;
      color: var(--muted);
      opacity: .9;
      text-align: center;
      padding: 0 12px;
      pointer-events: none;
    }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="top">
        <div>
          <h1>ГАРДЕРОБНАЯ СИСТЕМА ПРАКТИК HOME</h1>
          <div class="sub">PDF → CSV конвертер</div>
          <div class="hint">Загрузите PDF и скачайте CSV для импорта.</div>
        </div>
        <div class="badge">CSV: ; • UTF-8 • BOM</div>
      </div>

      <div class="row">
        <div class="file">
          <input id="pdf" type="file" accept="application/pdf,.pdf" />
        </div>
        <button id="btn" disabled>Скачать CSV</button>
      </div>

      <div id="status" class="status"></div>

      <div class="help" id="help" style="display:none;">
        <div class="helphead">
          <div class="helptitle">Мини-инструкция</div>
          <div class="openfull" id="openfull">Открыть полностью</div>
        </div>
        <div class="thumb" id="thumb">
          <video src="/instruction.mp4" muted playsinline preload="metadata"></video>
        </div>
      </div>
    </div>
  </div>

  <div class="corner" id="counter">…</div>
  <div class="footer-note">Программа создана командой ПРОМЕТ для своих дилеров</div>

  <div class="modal" id="modal" aria-hidden="true">
    <div class="modalcard">
      <div class="modalbar">
        <div class="t">Инструкция</div>
        <button class="close" id="close">Закрыть</button>
      </div>
      <div class="modalbody">
        <video src="/instruction.mp4" controls playsinline preload="auto"></video>
      </div>
    </div>
  </div>

  <script>
    const input = document.getElementById('pdf');
    const btn = document.getElementById('btn');
    const statusEl = document.getElementById('status');
    const help = document.getElementById('help');
    const thumb = document.getElementById('thumb');
    const openfull = document.getElementById('openfull');
    const modal = document.getElementById('modal');
    const closeBtn = document.getElementById('close');

    function ok(msg){ statusEl.className='status ok'; statusEl.textContent=msg; }
    function err(msg){ statusEl.className='status err'; statusEl.textContent=msg; }
    function neutral(msg){ statusEl.className='status'; statusEl.textContent=msg||''; }

    function openModal(){
      modal.classList.add('open');
      modal.setAttribute('aria-hidden','false');
      try { const v = modal.querySelector('video'); if (v) v.play().catch(()=>{}); } catch(e) {}
    }
    function closeModal(){
      modal.classList.remove('open');
      modal.setAttribute('aria-hidden','true');
      try { const v = modal.querySelector('video'); if (v) { v.pause(); v.currentTime = 0; } } catch(e) {}
    }

    thumb.addEventListener('click', openModal);
    openfull.addEventListener('click', openModal);
    closeBtn.addEventListener('click', closeModal);
    modal.addEventListener('click', (e) => { if (e.target === modal) closeModal(); });
    document.addEventListener('keydown', (e) => { if (e.key === 'Escape') closeModal(); });

    // show video help if file exists
    fetch('/instruction.mp4', { method: 'HEAD' }).then(r => {
      if (r.ok) {
        help.style.display = 'block';
        try { const tv = document.querySelector('#thumb video'); if (tv) tv.play().catch(()=>{}); } catch(e) {}
      }
    });

    async function loadCounter(){
      try {
        const r = await fetch('/stats');
        if (!r.ok) return;
        const j = await r.json();
        if (typeof j.conversions === 'number') {
          document.getElementById('counter').textContent = String(j.conversions);
        }
      } catch(e) {}
    }
    loadCounter();

    input.addEventListener('change', () => {
      const f = input.files && input.files[0];
      btn.disabled = !f;
      neutral(f ? ('Выбран файл: ' + f.name) : '');
    });

    btn.addEventListener('click', async () => {
      const f = input.files && input.files[0];
      if (!f) return;

      btn.disabled = true;
      const start = Date.now();

      let timer = setInterval(() => {
        const sec = Math.floor((Date.now() - start) / 1000);
        neutral('Обработка… прошло ' + sec + ' сек');
      }, 500);

      try {
        // 1) стартуем async job
        const fd = new FormData();
        fd.append('file', f);

        neutral('Загружаю PDF…');
        const r = await fetch('/extract_async', { method: 'POST', body: fd });
        if (!r.ok) {
          let text = await r.text();
          try { const j = JSON.parse(text); if (j.detail) text = String(j.detail); } catch(e) {}
          throw new Error(text || ('HTTP ' + r.status));
        }
        const data = await r.json();
        const job_id = data.job_id;

        // 2) опрашиваем статус
        while (true) {
          const s = await fetch('/job/' + job_id);
          if (!s.ok) {
            let text = await s.text();
            try { const j = JSON.parse(text); if (j.detail) text = String(j.detail); } catch(e) {}
            throw new Error(text || ('HTTP ' + s.status));
          }
          const j = await s.json();

          const sec = Math.floor((Date.now() - start) / 1000);
          let msg = (j.message || 'Обработка…') + ' • ' + sec + ' сек';

          if (j.total_pages && j.processed_pages) {
            msg += ' • страниц: ' + j.processed_pages + '/' + j.total_pages;
          }
          neutral(msg);

          if (j.status === 'done') {
            // 3) скачиваем результат
            const dl = await fetch('/job/' + job_id + '/download');
            if (!dl.ok) {
              let text = await dl.text();
              try { const jj = JSON.parse(text); if (jj.detail) text = String(jj.detail); } catch(e) {}
              throw new Error(text || ('HTTP ' + dl.status));
            }
            const blob = await dl.blob();

            const filename = (j.filename || ((f.name || 'items.pdf').replace(/\\.pdf$/i, '') + '.csv'));
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = filename;
            document.body.appendChild(a);
            a.click();
            a.remove();
            URL.revokeObjectURL(url);

            ok('Готово! Файл скачан: ' + filename);
            loadCounter();
            break;
          }

          if (j.status === 'error') {
            throw new Error(j.message || 'Ошибка обработки');
          }

          await new Promise(res => setTimeout(res, 600));
        }

      } catch (e) {
        err('Ошибка: ' + String(e.message || e));
      } finally {
        clearInterval(timer);
        btn.disabled = !(input.files && input.files[0]);
      }
    });
  </script>
</body>
</html>
"""


# -------------------------
# Endpoints
# -------------------------
@app.get("/stats")
def stats():
    return {"conversions": get_counter()}


@app.get("/health")
def health():
    # lightweight health
    try:
        _ensure_job_dir()
        job_files = len([x for x in os.listdir(JOB_DIR) if x.endswith(".json")])
    except Exception:
        job_files = -1

    return {
        "status": "ok",
        "article_map_size": len(ARTICLE_MAP),
        "article_map_status": ARTICLE_MAP_STATUS,
        "category_value": CATEGORY_VALUE,
        "instruction_video_exists": os.path.exists(INSTRUCTION_VIDEO_PATH),
        "conversions": get_counter(),
        "job_dir": JOB_DIR,
        "job_files": job_files,
    }


@app.api_route("/", methods=["GET", "HEAD"], response_class=HTMLResponse)
def home():
    return HOME_HTML


@app.api_route("/instruction.mp4", methods=["GET", "HEAD"])
def instruction_video():
    if not os.path.exists(INSTRUCTION_VIDEO_PATH):
        raise HTTPException(status_code=404, detail="instruction.mp4 not found")
    return FileResponse(INSTRUCTION_VIDEO_PATH, media_type="video/mp4")


# OLD endpoint (sync) — оставили для совместимости
@app.post("/extract")
async def extract(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Загрузите PDF файл (.pdf).")

    pdf_bytes = await file.read()

    try:
        rows, stats_ = parse_items(pdf_bytes)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Не удалось распарсить PDF: {e}")

    if not rows:
        raise HTTPException(
            status_code=422,
            detail=(
                "Не удалось найти позиции. Поддерживаемые якоря:\n"
                "1) в одной строке: price ₽ qty sum ₽\n"
                "2) в разных строках: price ₽ -> qty -> sum ₽\n"
                f"debug={stats_}"
            ),
        )

    csv_bytes = make_csv_excel_friendly(rows)
    increment_counter()
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="items.csv"'},
    )


# NEW async endpoints (progress)
@app.post("/extract_async")
async def extract_async(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Загрузите PDF файл (.pdf).")

    pdf_bytes = await file.read()
    original_filename = file.filename or "items.pdf"

    job_id = str(uuid4())
    _set_job(
        job_id,
        created_at=time.time(),
        status="processing",
        message="Старт обработки",
        processed_pages=0,
        total_pages=0,
        filename=(os.path.splitext(original_filename)[0] or "items") + ".csv",
    )

    def worker():
        try:
            _set_job(job_id, message="Читаю PDF…")
            rows, st = parse_items(pdf_bytes)

            _set_job(
                job_id,
                processed_pages=int(st.get("processed_pages", 0) or 0),
                total_pages=int(st.get("total_pages", 0) or 0),
                message="Формирую CSV…",
            )

            if not rows:
                _set_job(job_id, status="error", message="Не удалось найти позиции", stats=st)
                return

            csv_bytes = make_csv_excel_friendly(rows)
            _set_job_result(job_id, csv_bytes)
            increment_counter()

            _set_job(
                job_id,
                status="done",
                message="Готово",
                stats=st,
            )
        except Exception as e:
            _set_job(job_id, status="error", message=f"Ошибка: {e}")

    threading.Thread(target=worker, daemon=True).start()
    return {"job_id": job_id}


@app.get("/job/{job_id}")
def job_status(job_id: str):
    j = _get_job(job_id)
    if not j:
        raise HTTPException(status_code=404, detail="job not found")

    return {
        "status": j.get("status"),
        "message": j.get("message"),
        "processed_pages": int(j.get("processed_pages", 0) or 0),
        "total_pages": int(j.get("total_pages", 0) or 0),
        "stats": j.get("stats"),
        "filename": j.get("filename"),
    }


import urllib.parse

@app.get("/job/{job_id}/download")
def job_download(job_id: str):
    j = _get_job(job_id)
    if not j:
        raise HTTPException(status_code=404, detail="job not found")
    if j.get("status") != "done":
        raise HTTPException(status_code=409, detail="job not done")

    csv_bytes = _get_job_result(job_id)
    if not csv_bytes:
        raise HTTPException(status_code=404, detail="result not found")

    # безопасное имя файла
    filename_utf8 = j.get("filename", "items.csv")
    filename_ascii = re.sub(r'[^A-Za-z0-9_.-]+', '_', filename_utf8)

    quoted = urllib.parse.quote(filename_utf8)

    headers = {
        # fallback для старых клиентов
        "Content-Disposition": (
            f'attachment; filename="{filename_ascii}"; '
            f"filename*=UTF-8''{quoted}"
        )
    }

    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers=headers,
    )
